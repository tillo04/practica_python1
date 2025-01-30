import re, os, sys, requests
from pathlib import PurePath
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.files.file import File
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import win32com.client
import subprocess

# Parametros de configuracion
FOLDER_NAME = "prueba nequi"
FILE_NAME = "None"
FILE_NAME_PATTERN = "None"

# Rutas de archivos
RUTA_CARPETA = r"U:\python\extraccion-sharepoint\Almacenamiento"

# Configuracion de Sharepoint
SITE_URL = "https://grupouribegco.sharepoint.com/sites/msteams_2cef6a_342943"
SITE_NAME = "msteams_2cef6a_342943"
DOC_LIBRARY = "Documentos compartidos/"

cookies = {
    'FedAuth': '77u/PD94bWwgdmVyc2lvbj0iMS4wIiBlbmNvZGluZz0idXRmLTgiPz48U1A+VjE0LDBoLmZ8bWVtYmVyc2hpcHwxMDAzMjAwMzk3NWVkOWE2QGxpdmUuY29tLDAjLmZ8bWVtYmVyc2hpcHxwcmFjdGljYW50ZS5hbmFsaXRpY2FmaW5hbmNpZXJhMUBnY28uY29tLmNvLDEzMzgwNzUxMzI5MDAwMDAwMCwxMzM4MDcyNjAzOTAwMDAwMDAsMTMzODIzOTgxNDkyMDE2MzYwLDE3OS41MC43OS42Niw2Nyw0YjQ0ZDU2OC1iMTZkLTQ0NzMtYjRlMi00M2RhNmNkMmRlMDksLDc3ZWJhMGQ3LWI3YTMtNGI2ZC04YzBhLTQ0YzljOTEwNjg1YSxiNjZjNzhhMS1mMDBiLTcwMDAtYTY0My03MzdjODI2NDNlZGMsZDJjYzc5YTEtMzA2MS03MDAwLWE2NDMtNzk5ODc2ZTFlMTA1LCwwLDEzMzgxOTY5MDIwMTc4MTM3NCwxMzM4MjIyNDYyMDE3ODEzNzQsLCxleUo0YlhOZlkyTWlPaUpiWENKRFVERmNJbDBpTENKNGJYTmZjM050SWpvaU1TSXNJbkJ5WldabGNuSmxaRjkxYzJWeWJtRnRaU0k2SW5CeVlXTjBhV05oYm5SbExtRnVZV3hwZEdsallXWnBibUZ1WTJsbGNtRXhRR2RqYnk1amIyMHVZMjhpTENKMWRHa2lPaUpEVUZScVNWUkNiVFF3YVRCWloySlZhREpmWlVGQkluMD0sMjY1MDQ2Nzc0Mzk5OTk5OTk5OSwxMzM4MTU5NjkzMjAwMDAwMDAsY2YzMWU1ZWQtOGUyYS00NWUxLWI2NGItOTg5YjNhYzRlNGQyLCwsLCwsMTE1MjkyMTUwNDYwNjg0Njk3NiwsMTkyODc1LHVYZWhRSlBsZVZqTkNiYWtVaEdENkl5RlFRayxJUDZxS19vWVBWSUQ0VUtGLVBpNHNsdG9uZUUsRkxRQTcxK2tneStjQndGb3pqRXRITURNbk9DVFlWckxSMXdEQkdMUUlEdzEzanRmVVFuTUZqT2t6NlZsV21WMGkzcS96aXdqdXBNZjEwZXBmaEI4MVN0NjlscHR1ZDQxeVB5WGRqWEROUXBTY2NiYXF3dFY3TTMweWltZFFmdW9zc3gyQ1dRWVZHUXdTbU1TS1lsVEh1UGVXNUFhNGk5UGYxUDlDQjM2U1hFRFVGK2JPRlFQL042aktUdkMxZWFSM3BzVjVsaG05dzdZSENYNmRCNXlsMHo0RnhmL0xZME43WkpoTnp0aVNrbmRZR1hSTm03VVNjSStzaFZtc2ZySFloTmx2VGQxTWhEajlvZUg1TThKZUVxejg4ZU5UcTBuOGRoNWJzbkhIZ01TcjNEei9YK0Zmc1k0V1Q2SDNqU0swd1V3cnRmbSswaXlLeXlVT0hrVFZBPT08L1NQPg==',
    'rtFa': 'uaqN0/h2H3H/+YX5HczKaJLwzk+p2JBqCJH3ZLezp2AmNGI0NGQ1NjgtYjE2ZC00NDczLWI0ZTItNDNkYTZjZDJkZTA5IzEzMzgxNTk2OTM0NzI2NjIxMyNiNjZjNzhhMS1mMDBiLTcwMDAtYTY0My03MzdjODI2NDNlZGMjcHJhY3RpY2FudGUuYW5hbGl0aWNhZmluYW5jaWVyYTElNDBnY28uY29tLmNvIzE5Mjg3NSNiM0lRaDQ1a3lhZnh2ZkpUTlNQM1JzdW9PRk0jdEt1NmRzX0N6eEZyMF8zYVVJdDlWa2EwQlFZMFvPWwPGDx4HKOAi3Bo+f3G3azQkfA1co7kmmDodnBun4Wfpl3JMcrLGD275mwj5lYFeJ+7Z/tVDExuOccWUecFD992AXhXpIiUHRDLkx7yqjuQlNJ1Cs7Hdc3X0HipW/au8pOBpsFO9eA0M757DLUFG2P7MRoEtkJGIrXXAle57n+ylIzNJkzZxokklIhcbQBW1luzGr39+NDBjRbtyY6anPXokjI1DsAUqF04tXW0aPeYPZECY97Jx5ZXG3n7g15dJRR6ArDMUZU9LUI8/Xv08+Fb9etu+7sJci/J7v/aYMchVIgK8USaHG+8LruLYONdYum/D/3ab1+j7qpiDy+oAAAA='
}

session = requests.Session()
session.cookies.update(cookies)

class CookieAuthContext(AuthenticationContext):
    def authenticate_request(self, request_options):
        cookie_header = "; ".join([f"{name}={value}" for name, value in cookies.items()])
        request_options.set_header('Cookie', cookie_header)

ctx_auth = CookieAuthContext(SITE_URL)

class SharePoint:
    def _auth(self):
        conn = ClientContext(SITE_URL, ctx_auth)
        return conn
    
    def _get_files_list(self, folder_name):
        conn = self._auth()
        target_folder_url = f'{DOC_LIBRARY}/{folder_name}'
        root_folder = conn.web.get_folder_by_server_relative_url(target_folder_url)
        root_folder.expand(["Files", "Folders"]).get().execute_query()
        return root_folder.files
    
    def download_file(self, file_name, folder_name):
        conn = self._auth()
        file_url = f'/sites/{SITE_NAME}/{DOC_LIBRARY}/{folder_name}/{file_name}'
        file = File.open_binary(conn, file_url)
        return file

def generar_tabla_xlsx(ruta_xlsx):
    wb = load_workbook(ruta_xlsx)
    ws = wb.active
    min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row
    table_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
    tabla = Table(displayName="Table1", ref=table_range)
    estilo = TableStyleInfo(
        name="TableStyleLight13",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)
    wb.save(ruta_xlsx)

def save_file(file_n, file_obj):
    file_dir_path = PurePath(RUTA_CARPETA, file_n)
    if file_n.endswith('.csv'):
        temp_csv_path = file_dir_path.with_suffix('.csv')
        with open(temp_csv_path, 'wb') as f:
            f.write(file_obj.content)
        df = pd.read_csv(temp_csv_path, skiprows=9, delimiter=';')
        
        df["prueba combinar"] = df.iloc[:, 1].astype(str) + " " + df.iloc[:, 2].astype(str)
        df.iloc[:, 4] = pd.to_numeric(df.iloc[:, 4], errors='coerce')
        df.iloc[:, 5] = pd.to_numeric(df.iloc[:, 5], errors='coerce')

        df["suma de columnas comision e iva"] = df.iloc[:, 4] + df.iloc[:, 5]
        
        df["multiplicacion de columnas comision e iva"] = df.iloc[:, 4] * df.iloc[:, 5]
        
        df["resta de columnas comision e iva"] = df.iloc[:, 4] - df.iloc[:, 5]
        
        df.iloc[:, 16] = df.iloc[:, 16].astype(str)
        col_16_split = df.iloc[:, 16].str.split(" ", expand=True)
        col_16_split.columns = [f"prueba division columna{i+1}" for i in range(col_16_split.shape[1])]
        df = pd.concat([df, col_16_split], axis=1)
        
        xlsx_path = file_dir_path.with_suffix('.xlsx')
        df.to_excel(xlsx_path, index=False)
        generar_tabla_xlsx(xlsx_path)
        os.remove(temp_csv_path)
    else:
        with open(file_dir_path, 'wb') as f:
            f.write(file_obj.content)
    
def get_file(file_n, folder):
    file_obj = SharePoint().download_file(file_n, folder)
    save_file(file_n, file_obj)

def get_files(folder):
    files_list = SharePoint()._get_files_list(folder)
    for file in files_list:
        get_file(file.name, folder)
    
def get_files_by_pattern(keyword, folder):
    files_list = SharePoint()._get_files_list(folder)
    for file in files_list:
        if re.search(keyword, file.name):
            get_file(file.name, folder)

def enviar_correo(ruta_carpeta):
    outlook = win32com.client.Dispatch("Outlook.Application")
    outlook.GetNamespace("MAPI").GetDefaultFolder(6).Display()
    mail = outlook.CreateItem(0)
    mail.Subject = "PRUEBA ENVIO AUTOMATICO DE CORREOS"
    mail.Body = "mensaje de prueba"
    mail.To = ""
    mail.CC = "practicante.analiticafinanciera1@gco.com.co"
    for file_name in os.listdir(ruta_carpeta):
        file_path = os.path.join(ruta_carpeta, file_name)
        if os.path.isfile(file_path):
            mail.Attachments.Add(file_path)

    mail.Send()
    email_sent = True
    return email_sent

def main():
    if FILE_NAME != 'None':
        get_file(FILE_NAME, FOLDER_NAME)
    elif FILE_NAME_PATTERN != 'None':
        get_files_by_pattern(FILE_NAME_PATTERN, FOLDER_NAME)
    else:
        get_files(FOLDER_NAME)
    
    email_sent = enviar_correo(RUTA_CARPETA)

if __name__ == "__main__":
    main()